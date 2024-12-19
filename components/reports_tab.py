from PyQt5.QtWidgets import (QWidget, QLabel, QPushButton, QVBoxLayout, QHBoxLayout,
                             QTableWidget, QTableWidgetItem, QComboBox, QDateEdit, QMessageBox, QFileDialog)
from PyQt5.QtCore import Qt, QDate
import openpyxl
from openpyxl.styles import Alignment


class ReportsTab(QWidget):
    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.layout = QVBoxLayout(self)

        # Виджеты для генерации отчетов
        self.report_type_label = QLabel("Тип отчета:", self)
        self.report_type_combobox = QComboBox(self)
        self.report_type_combobox.addItems(["Продажи по дням", "Продажи по неделям", "Продажи по месяцам"])

        self.start_date_label = QLabel("Дата начала:", self)
        self.start_date_edit = QDateEdit(self)
        self.start_date_edit.setDate(QDate.currentDate())

        self.end_date_label = QLabel("Дата окончания:", self)
        self.end_date_edit = QDateEdit(self)
        self.end_date_edit.setDate(QDate.currentDate())

        self.generate_report_button = QPushButton("Сформировать отчет", self)
        self.generate_report_button.clicked.connect(self.generate_report)

        # Таблица для отображения отчетов
        self.report_table = QTableWidget(self)
        self.report_table.setColumnCount(6)
        self.report_table.setHorizontalHeaderLabels(
            ["ID Заказа", "Название блюда", "Количество", "Сумма", "Дата продажи", "Сотрудник"]
        )

        # Макет для полей ввода и кнопки
        input_layout = QHBoxLayout()
        input_layout.addWidget(self.report_type_label)
        input_layout.addWidget(self.report_type_combobox)
        input_layout.addWidget(self.start_date_label)
        input_layout.addWidget(self.start_date_edit)
        input_layout.addWidget(self.end_date_label)
        input_layout.addWidget(self.end_date_edit)
        input_layout.addWidget(self.generate_report_button)
        self.layout.addLayout(input_layout)

        self.layout.addWidget(self.report_table)

        self.setLayout(self.layout)

    def generate_report(self):
        report_type = self.report_type_combobox.currentText()
        start_date = self.start_date_edit.date().toString('yyyy-MM-dd')
        end_date = self.end_date_edit.date().toString('yyyy-MM-dd')

        print(f"Start date: {start_date}, End date: {end_date}")  # Для отладки

        try:
            # SQL-запрос для получения данных
            query = """
            SELECT 
                menu.name,
                SUM(order_dishes.quantity) AS quantity,
                SUM(menu.price * order_dishes.quantity) AS total_price,
                DATE(orders.order_time) AS order_date
            FROM 
                orders
            LEFT JOIN order_dishes ON orders.order_id = order_dishes.order_id
            LEFT JOIN menu ON order_dishes.dish_id = menu.menu_item_id
            WHERE orders.order_time BETWEEN %s AND %s
            GROUP BY menu.name, DATE(orders.order_time)
            ORDER BY orders.order_time
            """
            self.db_manager.cursor.execute(query, (start_date, end_date))
            report_data = self.db_manager.cursor.fetchall()

            print("Report Data:", report_data)  # Выводим данные для отладки

            if not report_data:
                QMessageBox.warning(self, "Предупреждение", "Нет данных для отображения в отчете.")
                return

            # Отображение данных в таблице интерфейса
            self.report_table.setRowCount(len(report_data))
            for row_idx, row in enumerate(report_data):
                for col_idx, value in enumerate(row):
                    self.report_table.setItem(row_idx, col_idx, QTableWidgetItem(str(value)))

            # Экспорт в Excel
            self.export_to_excel(report_data, report_type, start_date, end_date)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сгенерировать отчет: {e}")

    def export_to_excel(self, data, report_type, start_date, end_date):
        print("Exporting to Excel")  # Для отладки
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить отчет", "", "Excel Files (*.xlsx);;All Files (*)"
        )
        if not file_path:
            return  # Пользователь отменил выбор файла

        # Создание Excel файла
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Отчет по заказам"

        # Заголовок отчета
        sheet.merge_cells("A1:F1")
        sheet["A1"] = f"Отчет: {report_type} ({start_date} - {end_date})"
        sheet["A1"].alignment = Alignment(horizontal="center")

        # Заголовки столбцов
        headers = ["ID Заказа", "Название блюда", "Количество", "Сумма", "Дата продажи", "Сотрудник"]
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=2, column=col_num).value = header

        # Заполнение данными
        for row_num, row_data in enumerate(data, start=3):
            for col_num, value in enumerate(row_data, start=1):
                sheet.cell(row=row_num, column=col_num).value = value

        # Сохранение файла
        try:
            workbook.save(file_path)
            QMessageBox.information(self, "Успех", f"Отчет успешно сохранен: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить отчет: {e}")

